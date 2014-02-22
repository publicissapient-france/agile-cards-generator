import os
from openpyxl.cell import coordinate_from_string, column_index_from_string, get_column_letter

import openpyxl
import datetime
from copy import deepcopy
import csv

US_CARD_NAME = 'US'
DATA_SUFFIX = ' Data'
TEMPLATE_SUFFIX = ' Template'


def write_us_card(card, worksheet, starting_row, vertical_position=0, horizontal_position=0):

    def duplicate_cell_with_offset(cell, worksheet=None, row=0, column=0):

        def duplicate_style(cell, new_cell, worksheet):
            # used info on https://groups.google.com/forum/#!topic/openpyxl-users/s27khYlovwUfor line below
            worksheet._styles[new_cell.address] = cell.style

        def duplicate_cell_dimensions(cell, new_cell, worksheet):
            worksheet.row_dimensions[coordinate_from_string(new_cell.address)[1]] = \
                worksheet.row_dimensions[coordinate_from_string(cell.address)[1]]
            worksheet.column_dimensions[coordinate_from_string(new_cell.address)[0]] = \
                worksheet.column_dimensions[coordinate_from_string(cell.address)[0]]

        def duplicate_cell_merge_info(cell, new_cell, worksheet):
            for range_string in worksheet._merged_cells:
                if cell.address == range_string.split(':')[0]:

                    min_col, min_row, max_col, max_row = get_mins_maxs_from_range(range_string)
                    rows_in_range = max_row - min_row + 1
                    columns_in_range = max_col - min_col + 1
                    worksheet.merge_cells('%s:%s' % (new_cell.address,
                                                     new_cell.offset(row=rows_in_range - 1,
                                                                     column=columns_in_range - 1).address))

                    # For some reason need also to apply style to each of the merged cells
                    for r_offset in range(rows_in_range):
                        for c_offset in range(columns_in_range):
                            worksheet._styles[new_cell.offset(row=r_offset, column=c_offset).address] = cell.style

        if not worksheet:
            worksheet = cell.parent

        if row == 0 and column == 0:
            new_cell = cell
        else:
            new_cell = cell.offset(row=row, column=column)
            new_cell.value = cell.value
            duplicate_style(cell, new_cell, worksheet)
            duplicate_cell_dimensions(cell, new_cell, worksheet)
            duplicate_cell_merge_info(cell, new_cell, worksheet)
        return new_cell

    def create_card_cells(starting_row, column_offset, row_offset, worksheet):
        for row in range(USCard.ROW_HEIGHT):
            for column in range(USCard.COLUMN_WIDTH):
                my_cell = worksheet.cell(row=row + starting_row, column=column)
                duplicate_cell_with_offset(my_cell, row=row_offset,
                                           column=column_offset)


    def duplicate_cells_value(card, starting_row, column_offset, row_offset, worksheet):
        my_cell = worksheet.cell(row=starting_row + row_offset, column=column_offset)
        my_cell.value = card.mmf
        my_cell.offset(0, 2).value = card.feature
        my_cell.offset(0, 4).value = card.project
        my_cell.offset(1, 4).value = card.size
        my_cell.offset(2, 0).value = card.title
        my_cell.offset(3, 0).value = card.date_backlog
        my_cell.offset(3, 2).value = card.date_dev
        my_cell.offset(3, 4).value = card.date_done

    row_offset = vertical_position * (USCard.ROW_HEIGHT + 1)
    column_offset = horizontal_position * (USCard.COLUMN_WIDTH + 1)

    create_card_cells(starting_row, column_offset, row_offset, worksheet)
    duplicate_cells_value(card, starting_row, column_offset, row_offset, worksheet)


def write_us_cards(workbook, project_cards_data, card_worksheets_properties):


    def set_intermediate_columns_width(cards_per_row, my_worksheet):
        for i in range(1, cards_per_row):
            column_letter = get_column_letter(i * (USCard.COLUMN_WIDTH + 1))
            column_width = float(1)

            if column_letter in my_worksheet.column_dimensions:
                my_worksheet.column_dimensions[column_letter].width = column_width
            else:
                my_worksheet.column_dimensions[column_letter] = openpyxl.worksheet.ColumnDimension(width=column_width)


    def set_intermediate_rows_height(cards, cards_per_row, my_worksheet, row_height, starting_row):
        if len(cards) > cards_per_row:
            for i in list(range(1, len(list(range(cards_per_row, len(cards), cards_per_row))) + 1)):
                row_idx = i * (USCard.ROW_HEIGHT + 1) + starting_row

                my_worksheet.cell(row=row_idx - 1, column=0).value = ' '

                if row_idx in my_worksheet.row_dimensions:
                    my_worksheet.row_dimensions[row_idx].height = row_height
                else:
                    my_row_dimension = openpyxl.worksheet.RowDimension()
                    my_row_dimension.height = row_height
                    my_worksheet.row_dimensions[row_idx] = my_row_dimension


    my_worksheet = deepcopy(workbook.get_sheet_by_name(US_CARD_NAME + TEMPLATE_SUFFIX))
    my_worksheet.title = US_CARD_NAME
    workbook.add_sheet(my_worksheet)

    cards = project_cards_data.us_cards

    vertical_position = 0
    horizontal_position = 0
    cards_per_row = 2
    row_height = 5
    starting_row = card_worksheets_properties.us_properties.nb_settings_rows

    for card in cards:
        write_us_card(card, my_worksheet, starting_row, vertical_position, horizontal_position)
        horizontal_position += 1
        if horizontal_position == cards_per_row:
            horizontal_position = 0
            vertical_position += 1

    set_intermediate_columns_width(cards_per_row, my_worksheet)

    set_intermediate_rows_height(cards, cards_per_row, my_worksheet, row_height, starting_row)


class ProjectCardsData():

    def __init__(self, us_cards=None):
        self.us_cards = us_cards


class CardWorksheetsProperties():
    def __init__(self, us_properties):
        self.us_properties = us_properties


class CardWorksheetProperties():
    def __init__(self, nb_settings_rows = 0, lines_of_cards_per_page = 5):
        self.nb_settings_rows = nb_settings_rows
        self.lines_of_cards_per_page = lines_of_cards_per_page


class USCard():

    # TODO move these two to CardsWorksheetProperties or other better place
    ROW_HEIGHT = 4
    COLUMN_WIDTH = 6

    def __init__(self, mmf='MMF:', feature='Feature:', project='Projet:', size='Taille', title='Titre de la US',
                          date_backlog='Date backlog:', date_dev='Date dev:', date_done='Date done'):
        self.mmf = mmf
        self.feature = feature
        self.project = project
        self.size = size
        self.title = title
        self.date_backlog = date_backlog
        self.date_dev = date_dev
        self.date_done = date_done


def load_cards(workbook):
    us_cards = []

    header_row_handled = False
    for row in workbook.get_sheet_by_name(US_CARD_NAME + DATA_SUFFIX).rows:
        if not header_row_handled:
            header_row_handled = True
        else:
            new_card = USCard(mmf=row[0].value, feature=row[1].value, project=row[2].value, size=row[3].value, title=row[4].value, date_backlog=row[5].value, date_dev=row[6].value, date_done=row[7].value)
            us_cards.append(new_card)

    return ProjectCardsData(us_cards)

def extract_cards_worksheet_properties(workbook):
    my_worksheet = workbook.get_sheet_by_name(US_CARD_NAME + TEMPLATE_SUFFIX)
    us_properties = CardWorksheetProperties(nb_settings_rows=my_worksheet.cell(row=0, column=1).value,
                                             lines_of_cards_per_page=my_worksheet.cell(row=1, column=1).value)
    return CardWorksheetsProperties(us_properties)


def get_mins_maxs_from_range(range_string):
    min_col, min_row = coordinate_from_string(range_string.split(':')[0])
    max_col, max_row = coordinate_from_string(range_string.split(':')[1])
    min_col = column_index_from_string(min_col)
    max_col = column_index_from_string(max_col)
    return (min_col, min_row, max_col, max_row)


def setup_worksheet_page(my_workbook, us_worksheet_name, project_cards_data, card_worksheets_properties):

    def set_page_setup(my_us_worksheet):
        my_us_worksheet.page_setup.fitToPage = True
        my_us_worksheet.page_setup.fitToHeight = 0


    def set_page_margins(my_us_worksheet):
        my_us_worksheet.page_margins.left = 0.1
        my_us_worksheet.page_margins.right = 0.1
        my_us_worksheet.page_margins.top = 0.1
        my_us_worksheet.page_margins.bottom = 0.1


    def add_page_breaks(my_worksheet, project_cards, card_worksheets_properties):
        nb_settings_rows = card_worksheets_properties.us_properties.nb_settings_rows
        lines_of_cards_per_page = card_worksheets_properties.us_properties.lines_of_cards_per_page

        lines_of_card = len(list(range(2, len(project_cards.us_cards), 2))) + 1
        nb_of_page_breaks = len(list(range(lines_of_cards_per_page, lines_of_card, lines_of_cards_per_page)))

        for i in range(1, nb_of_page_breaks + 1):
            page_break_row_idx = i * ((lines_of_cards_per_page + 1) * USCard.ROW_HEIGHT + 2) + nb_settings_rows
            my_worksheet.page_breaks.append(page_break_row_idx)


    def hide_settings_rows(card_worksheets_properties, my_us_worksheet):
        rows_idx = range(card_worksheets_properties.us_properties.nb_settings_rows)
        for row_idx in rows_idx:
            row_first_cell = my_us_worksheet.cell(row=row_idx, column=0)
            if not row_first_cell.value:
                row_first_cell.value = ' '
            my_us_worksheet.row_dimensions[row_idx + 1].visible = False


    my_us_worksheet = my_workbook.get_sheet_by_name(us_worksheet_name)
    my_workbook.active = my_workbook.get_index(my_us_worksheet)

    set_page_setup(my_us_worksheet)
    set_page_margins(my_us_worksheet)
    add_page_breaks(my_us_worksheet, project_cards_data, card_worksheets_properties)
    hide_settings_rows(card_worksheets_properties, my_us_worksheet)


def main():
    output_file_name = prepare_output_file(None, 'xlsx')
    input_file_name = os.path.join('input', 'input.xlsx')

    my_workbook = openpyxl.load_workbook(input_file_name)

    project_cards_data = load_cards(my_workbook)

    card_worksheets_properties = extract_cards_worksheet_properties(my_workbook)

    write_us_cards(my_workbook, project_cards_data, card_worksheets_properties)

    setup_worksheet_page(my_workbook, US_CARD_NAME, project_cards_data, card_worksheets_properties)

    my_workbook.save(output_file_name)

    text = 'Time at which file was generated: '+ datetime.datetime.now().__str__()

    print(text)


def prepare_output_file(output_file, extension):
    file_name = None
    if output_file is not None:
        file_name = output_file
    else:
        file_name = 'output.' + extension

    output_path = 'output'
    if not os.path.isdir(output_path):
        os.makedirs(output_path, exist_ok=True)
    file_name = os.path.join(output_path, file_name)

    if os.path.isfile(file_name):
        os.remove(file_name)

    return file_name


if __name__ == "__main__":
    main()